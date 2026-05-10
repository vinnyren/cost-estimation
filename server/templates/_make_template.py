"""一次性运行：生成 report-v1.xlsx 模板。"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).parent / "report-v1.xlsx"


def _bold() -> Font:
    return Font(bold=True)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_template() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 1. 封面声明
    ws = wb.create_sheet("封面声明")
    ws["A1"] = "第三方软件造价评估报告"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A3"] = "项目名称："
    ws["A4"] = "委托单位："
    ws["A5"] = "评估内容："
    ws["A6"] = "评估编号："
    ws["A7"] = "报告日期："

    # 2. 评估结果摘要
    ws = wb.create_sheet("评估结果摘要")
    ws.append(["序号", "评估内容", "评估结果", "单位"])
    for c in ws[1]:
        c.font = _bold()
        c.alignment = _center()
    ws.append([1, "功能点规模", None, "FP"])
    ws.append([2, "功能点开发工作量（下限）", None, "人时"])
    ws.append([2, "功能点开发工作量（中值）", None, "人时"])
    ws.append([2, "功能点开发工作量（上限）", None, "人时"])
    ws.append([3, "软件开发费用（下限）", None, "万元"])
    ws.append([3, "软件开发费用（中值）", None, "万元"])
    ws.append([3, "软件开发费用（上限）", None, "万元"])
    ws.append([4, "总费用（中值）", None, "万元"])
    # 命名区域便于 exporter 填值
    wb.defined_names["scale_adjusted"] = DefinedName("scale_adjusted", attr_text="评估结果摘要!$C$2")
    wb.defined_names["effort_dev_p10"] = DefinedName("effort_dev_p10", attr_text="评估结果摘要!$C$3")
    wb.defined_names["effort_dev_p50"] = DefinedName("effort_dev_p50", attr_text="评估结果摘要!$C$4")
    wb.defined_names["effort_dev_p90"] = DefinedName("effort_dev_p90", attr_text="评估结果摘要!$C$5")
    wb.defined_names["cost_dev_p10"] = DefinedName("cost_dev_p10", attr_text="评估结果摘要!$C$6")
    wb.defined_names["cost_dev_p50"] = DefinedName("cost_dev_p50", attr_text="评估结果摘要!$C$7")
    wb.defined_names["cost_dev_p90"] = DefinedName("cost_dev_p90", attr_text="评估结果摘要!$C$8")
    wb.defined_names["cost_total_p50"] = DefinedName("cost_total_p50", attr_text="评估结果摘要!$C$9")

    # 3. 评估报告书
    ws = wb.create_sheet("评估报告书")
    ws["A1"] = "一、项目概述"
    ws["A1"].font = _bold()
    ws["A3"] = "二、评估目的"
    ws["A3"].font = _bold()
    ws["A5"] = "三、评估依据/技术/方法"
    ws["A5"].font = _bold()
    wb.defined_names["project_overview"] = DefinedName("project_overview", attr_text="评估报告书!$A$2")
    wb.defined_names["evaluation_purpose"] = DefinedName("evaluation_purpose", attr_text="评估报告书!$A$4")

    # 4. 调整因子表
    ws = wb.create_sheet("调整因子表")
    ws.append(["类别", "调整因子", "取值", "说明"])
    for c in ws[1]:
        c.font = _bold()
        c.alignment = _center()

    # 5. 功能点计数表
    ws = wb.create_sheet("功能点计数表")
    ws.append(["编号", "子系统", "一级模块", "二级模块", "功能项描述",
               "功能点计数项名称", "类别", "UFP", "重用程度", "修改类型", "US", "来源", "备注"])
    for c in ws[1]:
        c.font = _bold()
        c.alignment = _center()

    # 6. 详细计算过程
    ws = wb.create_sheet("详细计算过程")
    ws.append(["步骤", "说明", "公式", "结果"])
    for c in ws[1]:
        c.font = _bold()
        c.alignment = _center()

    # 7. 参数附录
    ws = wb.create_sheet("参数附录")
    ws.append(["参数名", "取值", "来源", "备注"])
    for c in ws[1]:
        c.font = _bold()
        c.alignment = _center()

    wb.save(OUT)
    print(f"✓ template written: {OUT}")


if __name__ == "__main__":
    make_template()
