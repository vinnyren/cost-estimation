from dataclasses import dataclass, field
from pathlib import Path
from openpyxl import load_workbook


@dataclass
class ParsedSheet:
    name: str
    headers: list[str]
    rows: list[list]
    metadata: dict = field(default_factory=dict)


def parse_xlsx(path: Path) -> list[ParsedSheet]:
    """提取 Excel 全部 Sheet。第一行为 header；空表跳过"""
    if not path.exists():
        raise FileNotFoundError(f"XLSX 文件不存在: {path}")
    wb = load_workbook(str(path), data_only=True, read_only=True)
    out: list[ParsedSheet] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(c) if c is not None else "" for c in header_row]
        rows: list[list] = []
        for row in rows_iter:
            if all(c is None for c in row):
                continue
            rows.append([c if c is not None else "" for c in row])
        out.append(ParsedSheet(name=ws.title, headers=headers, rows=rows,
                                 metadata={"source": str(path), "type": "xlsx"}))
    wb.close()
    return out
