"""模板损坏时的内置生成器：用代码直接构建一个最小可用的 7-Sheet 报告。"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def _bold() -> Font:
    return Font(bold=True)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def render_fallback(output_path: Path, *,
                    project_name: str, project_overview: str,
                    scale_adjusted: float, effort_dev: dict, cost_dev: dict,
                    cost_total_p50_yuan: float,
                    functions: list[dict], factors: list[dict],
                    steps: list[dict], params: list[dict]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("封面声明")
    cover["A1"] = "第三方软件造价评估报告（fallback 版）"
    cover["A1"].font = Font(bold=True, size=18)
    cover["A3"] = f"项目名称：{project_name}"

    summary = wb.create_sheet("评估结果摘要")
    summary.append(["项目", "数值", "单位"])
    for c in summary[1]:
        c.font = _bold()
        c.alignment = _center()
    summary.append(["调整后规模", round(scale_adjusted, 2), "FP"])
    summary.append(["开发工作量 P50", round(effort_dev["P50"], 2), "人时"])
    summary.append(["开发成本 P50", round(cost_dev["P50"] / 10000, 4), "万元"])
    summary.append(["总费用 P50", round(cost_total_p50_yuan / 10000, 4), "万元"])

    report = wb.create_sheet("评估报告书")
    report["A1"] = "项目概述"
    report["A1"].font = _bold()
    report["A2"] = project_overview

    factors_ws = wb.create_sheet("调整因子表")
    factors_ws.append(["类别", "名称", "取值"])
    for f in factors:
        factors_ws.append([f.get("category", ""), f.get("name", ""), f.get("value", 0)])

    fp_ws = wb.create_sheet("功能点计数表")
    fp_ws.append(["编号", "名称", "类别", "UFP", "US"])
    for i, fp in enumerate(functions, start=1):
        fp_ws.append([i, fp.get("name", ""), fp.get("category", ""), fp.get("ufp", 0), fp.get("us", 0)])

    steps_ws = wb.create_sheet("详细计算过程")
    steps_ws.append(["步骤", "说明", "公式", "结果"])
    for s in steps:
        steps_ws.append([s.get("step", ""), s.get("desc", ""), s.get("formula", ""), s.get("result", "")])

    params_ws = wb.create_sheet("参数附录")
    params_ws.append(["键", "值"])
    for p in params:
        params_ws.append([p.get("key", ""), p.get("value", "")])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
