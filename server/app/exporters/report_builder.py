"""软件造价评估报告 Excel 生成器（参考行业评估表格式）。

产出 6 个 Sheet：
  1. 封面          —— 报告标题 + 项目信息 + 声明
  2. 评估结果汇总   —— 规模 / 工作量 / 成本 / 造价主表（三档）
  3. 模块功能点及费用分项统计表 —— 按一级模块分摊
  4. 系统功能点明细表 —— 逐条功能点
  5. 评估报告书     —— 文字版评估报告（项目概述 / 依据 / 方法 / 结论 / 说明）
  6. 调整因子表     —— 本次评估采用的因子取值

旧版用二进制模板 report-v1.xlsx + 命名单元格，模板的「评估报告书」表
无命名单元格因而永远空白。本模块全代码构建，不依赖模板。
"""
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .excel import _safe_text

_BANDS = ("P10", "P50", "P90")
_BAND_LABEL = {"P10": "下限值", "P50": "中值", "P90": "上限值"}

_TITLE_FONT = Font(bold=True, size=16)
_SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEAD_FONT = Font(bold=True, size=10)
_BODY_FONT = Font(size=10)
_SECTION_FILL = PatternFill("solid", fgColor="165DFF")
_HEAD_FILL = PatternFill("solid", fgColor="E8EEF9")
_THIN = Side(style="thin", color="D0D5DD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _money(v: float) -> float:
    return round(float(v or 0.0), 2)


def _wan(v: float) -> float:
    return round(float(v or 0.0) / 10000.0, 4)


def build_report(
    out_path: Path,
    *,
    project: Any,
    functions: list,
    figures: dict,
    is_reverse: bool,
    target_cost_wan: float | None,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_cover(wb, project, is_reverse, target_cost_wan)
    _sheet_summary(wb, project, functions, figures, is_reverse, target_cost_wan)
    _sheet_modules(wb, functions, figures)
    _sheet_fp_detail(wb, functions)
    _sheet_narrative(wb, project, functions, figures, is_reverse, target_cost_wan)
    _sheet_factors(wb, figures)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------- 封面
def _sheet_cover(wb: Workbook, project: Any, is_reverse: bool,
                 target_cost_wan: float | None) -> None:
    ws = wb.create_sheet("封面")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 58
    ws["A1"] = "软件造价评估报告"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = _CENTER

    mode_label = "反算（目标造价 → 反推规模）" if is_reverse else "正向（功能点 → 造价）"
    rows = [
        ("项目名称", project.name),
        ("项目编号", project.id),
        ("委托单位", project.client or "—"),
        ("评估单位", project.evaluator or "—"),
        ("评估阶段", project.phase),
        ("所在城市", project.city),
        ("所属行业", project.industry),
        ("评估方式", mode_label),
        ("基准数据版本", project.basis_data_ver),
        ("出具日期", date.today().isoformat()),
    ]
    if is_reverse and target_cost_wan is not None:
        rows.insert(7, ("目标造价", f"{target_cost_wan:g} 万元"))

    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = _HEAD_FONT
        ws.cell(r, 1).fill = _HEAD_FILL
        ws.cell(r, 1).border = _BORDER
        c = ws.cell(r, 2, _safe_text(v))
        c.font = _BODY_FONT
        c.border = _BORDER
        c.alignment = _WRAP
        r += 1

    r += 1
    ws.cell(r, 1, "声明").font = _HEAD_FONT
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)
    note = ws.cell(
        r + 1, 1,
        "本报告依据《软件研发成本度量规范》（GB/T 36964-2018）及 CSBMK 基准数据，"
        "采用 NESMA 功能点方法测算。结果供造价参考，最终以合同约定为准。",
    )
    note.font = _BODY_FONT
    note.alignment = _WRAP
    ws.row_dimensions[r + 1].height = 48


# ---------------------------------------------------------------- 评估结果汇总
def _sheet_summary(wb: Workbook, project: Any, functions: list, fig: dict,
                   is_reverse: bool, target_cost_wan: float | None) -> None:
    ws = wb.create_sheet("评估结果汇总")
    for col, w in zip("ABCDE", (16, 26, 12, 20, 16)):
        ws.column_dimensions[col].width = w

    ws["A1"] = "评估结果汇总表"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = _CENTER

    head = ["类别", "评估指标", "档位", "评估结果", "单位"]
    for i, h in enumerate(head, 1):
        c = ws.cell(3, i, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    ufp_total = sum(float(fp.ufp or 0) for fp in functions)
    s_adj = fig["scale_adjusted"]
    dev_factor = fig["dev_factor"] or 1.0
    eff = fig["effort_dev"]
    cost_dev = fig["cost_dev"]
    cost_ops = fig.get("cost_ops", {})
    cost_total_p50 = fig["cost_total_p50_yuan"]
    # 基准生产率（不含调整因子）：AE = S × pdr × 因子 → pdr = AE /(S×因子)。
    # 反算每档 S 不同，按档取对应规模。
    s_bands = fig.get("scale_adjusted_bands", {b: s_adj for b in _BANDS})
    pdr = {
        b: (eff[b] / (s_bands[b] * dev_factor)
            if s_bands[b] * dev_factor > 0 else 0.0)
        for b in _BANDS
    }

    r = 4

    def section(name: str) -> None:
        nonlocal r
        c = ws.cell(r, 1, name)
        c.font = _SECTION_FONT
        c.fill = _SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c.alignment = Alignment(vertical="center")
        for col in range(1, 6):
            ws.cell(r, col).border = _BORDER
        r += 1

    def row(cat: str, metric: str, band: str, value: Any, unit: str) -> None:
        nonlocal r
        vals = [cat, metric, band, value, unit]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, _safe_text(v))
            c.font = _BODY_FONT
            c.border = _BORDER
            c.alignment = _CENTER if i in (3, 5) else _WRAP
        r += 1

    section("一、软件规模度量")
    row("规模", "未调整功能点数 UFP", "—", round(ufp_total, 2), "功能点")
    row("规模", "未调整规模 US", "—", round(fig["scale_us"], 2), "功能点")
    row("规模", "规模变更因子 CF", "—", round(fig["cf_used"], 4), "—")
    row("规模", "调整后规模 S", "—", round(s_adj, 2), "功能点")

    section("二、工作量度量")
    for b in _BANDS:
        row("工作量", "基准生产率 PDR", _BAND_LABEL[b], round(pdr[b], 4), "人时/功能点")
    row("工作量", "综合调整因子", "—", round(dev_factor, 4), "—")
    for b in _BANDS:
        row("工作量", "调整后工作量 AE", _BAND_LABEL[b], round(eff[b], 2), "人时")

    section("三、成本估算")
    row("成本", "人月折算系数", "—", round(fig["hours_per_pm"], 2), "人时/人月")
    row("成本", "人力成本费率", "—", round(fig["rate_dev"], 2), "元/人月")
    for b in _BANDS:
        row("成本", "软件开发费用 P", _BAND_LABEL[b], _money(cost_dev[b]), "元")
    ops_p50 = float(cost_ops.get("P50", 0.0) or 0.0)
    if ops_p50 > 0:
        row("成本", "运维费用", "中值", _money(ops_p50), "元")
    row("成本", "其他费用 DNC", "—", _money(fig.get("other_cost", 0.0)), "元")

    section("四、评估结论")
    row("结论", "评估总造价", "中值", _money(cost_total_p50), "元")
    row("结论", "评估总造价", "中值", _wan(cost_total_p50), "万元")
    unit_price = cost_total_p50 / ufp_total if ufp_total > 0 else 0.0
    row("结论", "功能点单价", "—", round(unit_price, 2), "元/功能点")
    if is_reverse and target_cost_wan is not None:
        row("结论", "目标造价（反算输入）", "—", round(target_cost_wan, 4), "万元")


# ------------------------------------------ 模块功能点及费用分项统计表
def _sheet_modules(wb: Workbook, functions: list, fig: dict) -> None:
    ws = wb.create_sheet("模块功能点及费用分项统计表")
    for col, w in zip("ABCDEFG", (8, 16, 22, 16, 16, 18, 12)):
        ws.column_dimensions[col].width = w

    ws["A1"] = "系统模块功能点、工作量及费用分项统计表"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = _CENTER

    head = ["序号", "子系统", "一级模块", "功能点数 UFP",
            "调整后工作量(人日)", "开发费用(元)", "费用占比"]
    for i, h in enumerate(head, 1):
        c = ws.cell(3, i, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    # 按 (子系统, 一级模块) 分组累计
    groups: dict[tuple, dict] = {}
    for fp in functions:
        key = (fp.subsystem or "—", fp.l1_module or "未分类")
        g = groups.setdefault(key, {"ufp": 0.0, "us": 0.0})
        g["ufp"] += float(fp.ufp or 0)
        g["us"] += float(fp.us or 0)

    total_us = sum(g["us"] for g in groups.values()) or 1.0
    total_cost = fig["cost_total_p50_yuan"]
    total_effort_day = fig["effort_dev"]["P50"] / 8.0

    r = 4
    for idx, ((sub, mod), g) in enumerate(sorted(groups.items()), 1):
        ratio = g["us"] / total_us
        vals = [
            idx, sub, mod, round(g["ufp"], 2),
            round(total_effort_day * ratio, 2),
            _money(total_cost * ratio),
            f"{ratio * 100:.2f}%",
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, _safe_text(v))
            c.font = _BODY_FONT
            c.border = _BORDER
            c.alignment = _CENTER if i in (1, 4, 5, 6, 7) else _WRAP
        r += 1

    # 汇总行
    sum_vals = ["", "项目汇总", "",
                round(sum(g["ufp"] for g in groups.values()), 2),
                round(total_effort_day, 2), _money(total_cost), "100.00%"]
    for i, v in enumerate(sum_vals, 1):
        c = ws.cell(r, i, v)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _BORDER
        c.alignment = _CENTER if i in (1, 4, 5, 6, 7) else _WRAP


# ---------------------------------------------------------------- 功能点明细
def _sheet_fp_detail(wb: Workbook, functions: list) -> None:
    ws = wb.create_sheet("系统功能点明细表")
    widths = (8, 14, 16, 18, 12, 22, 28, 8, 10, 10, 10, 12, 20)
    for col, w in zip("ABCDEFGHIJKLM", widths):
        ws.column_dimensions[col].width = w

    ws["A1"] = "系统功能点明细表"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:M1")
    ws["A1"].alignment = _CENTER

    head = ["编号", "子系统", "一级模块", "二级模块", "类别",
            "功能点计数项名称", "功能项描述", "UFP", "重用程度",
            "修改类型", "US", "来源", "备注"]
    for i, h in enumerate(head, 1):
        c = ws.cell(3, i, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    r = 4
    for idx, fp in enumerate(functions, 1):
        vals = [
            idx, fp.subsystem, fp.l1_module, fp.l2_module, fp.category,
            fp.name, fp.description, fp.ufp, fp.reuse_level, fp.modify_type,
            fp.us, fp.source, fp.notes,
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, _safe_text(v))
            c.font = _BODY_FONT
            c.border = _BORDER
            c.alignment = _CENTER if i in (1, 5, 8, 9, 10, 11) else _WRAP
        r += 1


# ---------------------------------------------------------------- 评估报告书
def _sheet_narrative(wb: Workbook, project: Any, functions: list, fig: dict,
                     is_reverse: bool, target_cost_wan: float | None) -> None:
    ws = wb.create_sheet("评估报告书")
    ws.column_dimensions["A"].width = 110

    ws["A1"] = "软件造价评估报告书"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER

    ufp_total = sum(float(fp.ufp or 0) for fp in functions)
    p50 = fig["cost_total_p50_yuan"]
    p10 = fig["cost_total"]["P10"]
    p90 = fig["cost_total"]["P90"]
    mode_label = "反算" if is_reverse else "正向"

    paras: list[tuple[str, str]] = [
        ("一、项目概述",
         f"本次评估对象为「{project.name}」（项目编号 {project.id}）。"
         f"委托单位 {project.client or '—'}，评估单位 {project.evaluator or '—'}，"
         f"所在城市 {project.city}，所属行业 {project.industry}，"
         f"评估阶段为「{project.phase}」，评估方式为{mode_label}。"),
        ("二、评估依据",
         "1.《软件研发成本度量规范》（GB/T 36964-2018）；"
         "2. NESMA 功能点估算方法；"
         f"3. CSBMK 行业基准数据（版本 {project.basis_data_ver}）。"),
        ("三、评估方法",
         "采用功能点方法测算软件规模：先统计未调整功能点数 UFP，乘以规模变更"
         f"因子 CF={round(fig['cf_used'], 3)} 得到调整后规模 S；再以行业基准生产率"
         "结合综合调整因子折算工作量，最后按人力成本费率换算为开发费用。"
         "结果按 P10/P50/P90 三档给出，推荐采用 P50 中位档。"),
        ("四、评估过程",
         f"经测算，本项目未调整功能点数 UFP 为 {round(ufp_total, 2)} 功能点，"
         f"调整后规模 S 为 {round(fig['scale_adjusted'], 2)} 功能点；"
         f"调整后工作量（中值）{round(fig['effort_dev']['P50'], 2)} 人时；"
         f"人力成本费率 {round(fig['rate_dev'], 2)} 元/人月。"
         + (f" 本项目为反算评估，目标造价 {target_cost_wan:g} 万元，"
            "据此反推可承载的功能点规模。" if is_reverse and target_cost_wan
            else "")),
        ("五、评估结论",
         f"本项目评估总造价区间为 {_money(p10):,.2f} 元 ~ {_money(p90):,.2f} 元，"
         f"推荐采用中位值 P50：{_money(p50):,.2f} 元（约 {_wan(p50):g} 万元）。"),
        ("六、费用说明",
         "上述费用为软件开发费用测算结果，未包含直接非人力费用（DNC）、"
         "硬件采购、第三方软件授权等。运维费用如有需要按相应口径单独测算。"
         "本报告结果供造价参考，最终以合同约定为准。"),
    ]

    r = 3
    for title, body in paras:
        t = ws.cell(r, 1, title)
        t.font = _HEAD_FONT
        r += 1
        b = ws.cell(r, 1, body)
        b.font = _BODY_FONT
        b.alignment = _WRAP
        ws.row_dimensions[r].height = 60
        r += 2


# ---------------------------------------------------------------- 调整因子表
def _sheet_factors(wb: Workbook, fig: dict) -> None:
    ws = wb.create_sheet("调整因子表")
    for col, w in zip("ABC", (28, 18, 56)):
        ws.column_dimensions[col].width = w

    ws["A1"] = "本次评估采用的调整因子"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = _CENTER

    head = ["因子", "取值", "说明"]
    for i, h in enumerate(head, 1):
        c = ws.cell(3, i, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    rows = [
        ("规模变更因子 CF", round(fig["cf_used"], 4),
         "按评估阶段取值（早期 1.39 / 中期 1.21 / 晚期 1.10 / 运维 1.00）"),
        ("综合调整因子（开发）", round(fig["dev_factor"], 4),
         "非功能性特征、团队背景、完整性级别、开发平台、应用类型等因子之积"),
        ("人月折算系数", round(fig["hours_per_pm"], 2),
         "人时/人月，按 21.75 天 × 8 人时/天 = 174 人时取值"),
        ("人力成本费率", round(fig["rate_dev"], 2),
         "元/人月，按城市基准费率取值"),
    ]
    r = 4
    for k, v, desc in rows:
        for i, val in enumerate((k, v, desc), 1):
            c = ws.cell(r, i, _safe_text(val))
            c.font = _BODY_FONT
            c.border = _BORDER
            c.alignment = _CENTER if i == 2 else _WRAP
        r += 1
