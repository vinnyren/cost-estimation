from datetime import datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

from ..config import settings


REQUIRED_SHEETS = ["封面声明", "评估结果摘要", "评估报告书", "调整因子表",
                   "功能点计数表", "详细计算过程", "参数附录"]
REQUIRED_NAMES = ["scale_adjusted", "effort_dev_p10", "effort_dev_p50", "effort_dev_p90",
                  "cost_dev_p10", "cost_dev_p50", "cost_dev_p90", "cost_total_p50"]


# OWASP CSV / Excel injection 防护：用户输入字符串若以这些字符开头，Excel
# 会把它当公式执行（=HYPERLINK / =cmd|... 等是已知 RCE 路径）。在写入前
# 加单引号让 Excel 显式当文本处理。
# 覆盖范围：ASCII = + - @ + Tab + CR。Excel 不把全角 ＝＋－＠ 视为公式触发，
# 所以不需要列入。NUL/换行不会被 Excel 解析为公式。
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_text(v: Any) -> Any:
    """If v is a string starting with a formula-trigger char, prefix it with
    a single quote so Excel renders it as text. Pass through everything else
    (numbers, None) as-is."""
    if isinstance(v, str) and v.startswith(_FORMULA_TRIGGERS):
        return "'" + v
    return v


class TemplateBrokenError(RuntimeError):
    pass


def _validate_template(wb) -> None:
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise TemplateBrokenError(f"missing sheets: {missing_sheets}")
    names = set(wb.defined_names.keys())
    missing_names = [n for n in REQUIRED_NAMES if n not in names]
    if missing_names:
        raise TemplateBrokenError(f"missing named ranges: {missing_names}")


def _write_named(wb, name: str, value) -> None:
    dn: DefinedName = wb.defined_names[name]
    coord = dn.attr_text.split("!")[1].replace("$", "")
    sheet_name = dn.attr_text.split("!")[0].strip("'")
    wb[sheet_name][coord] = value


def render(template_path: Path, output_path: Path, *,
           project_name: str, project_overview: str,
           scale_adjusted: float, effort_dev: dict, cost_dev: dict, cost_total_p50_yuan: float,
           functions: list[dict], factors: list[dict], steps: list[dict], params: list[dict]) -> Path:
    """加载模板，按命名区域填值，导出"""
    wb = load_workbook(str(template_path))
    _validate_template(wb)

    # 摘要数值
    _write_named(wb, "scale_adjusted", round(scale_adjusted, 2))
    _write_named(wb, "effort_dev_p10", round(effort_dev["P10"], 2))
    _write_named(wb, "effort_dev_p50", round(effort_dev["P50"], 2))
    _write_named(wb, "effort_dev_p90", round(effort_dev["P90"], 2))
    _write_named(wb, "cost_dev_p10", round(cost_dev["P10"] / 10000, 4))
    _write_named(wb, "cost_dev_p50", round(cost_dev["P50"] / 10000, 4))
    _write_named(wb, "cost_dev_p90", round(cost_dev["P90"] / 10000, 4))
    _write_named(wb, "cost_total_p50", round(cost_total_p50_yuan / 10000, 4))

    # 报告书
    _write_named(wb, "project_overview", project_overview or "")

    # 封面（直接写单元格）— 项目名 用 _safe_text 包一道（前缀已经是「项目
    # 名称：」所以实际上不会触发 _FORMULA_TRIGGERS，但保持纵深防御）
    cover = wb["封面声明"]
    cover["A3"] = _safe_text(f"项目名称：{project_name}")
    cover["A7"] = f"报告日期：{datetime.now().strftime('%Y-%m-%d')}"

    # FP 计数表 — 所有用户输入字符串都过 _safe_text
    ws = wb["功能点计数表"]
    for i, fp in enumerate(functions, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, _safe_text(fp.get("subsystem", "")))
        ws.cell(i, 3, _safe_text(fp.get("l1_module", "")))
        ws.cell(i, 4, _safe_text(fp.get("l2_module", "")))
        ws.cell(i, 5, _safe_text(fp.get("description", "")))
        ws.cell(i, 6, _safe_text(fp.get("name", "")))
        ws.cell(i, 7, _safe_text(fp.get("category", "")))
        ws.cell(i, 8, fp.get("ufp", 0))
        ws.cell(i, 9, _safe_text(fp.get("reuse_level", "")))
        ws.cell(i, 10, _safe_text(fp.get("modify_type", "")))
        ws.cell(i, 11, fp.get("us", 0))
        ws.cell(i, 12, _safe_text(fp.get("source", "")))
        ws.cell(i, 13, _safe_text(fp.get("notes", "")))

    # 调整因子
    ws = wb["调整因子表"]
    for i, f in enumerate(factors, start=2):
        ws.cell(i, 1, _safe_text(f.get("category", "")))
        ws.cell(i, 2, _safe_text(f.get("name", "")))
        ws.cell(i, 3, f.get("value", 0))
        ws.cell(i, 4, _safe_text(f.get("note", "")))

    # 详细计算过程 — 注意 formula 字段是真公式描述（如「Σ us」），数值结果
    # 用 _safe_text 没意义；其它文本字段还是要包
    ws = wb["详细计算过程"]
    for i, s in enumerate(steps, start=2):
        ws.cell(i, 1, _safe_text(s.get("step", "")))
        ws.cell(i, 2, _safe_text(s.get("desc", "")))
        ws.cell(i, 3, _safe_text(s.get("formula", "")))
        ws.cell(i, 4, s.get("result", ""))

    # 参数附录
    ws = wb["参数附录"]
    for i, p in enumerate(params, start=2):
        ws.cell(i, 1, _safe_text(p.get("key", "")))
        ws.cell(i, 2, _safe_text(p.get("value", "")))
        ws.cell(i, 3, _safe_text(p.get("source", "")))
        ws.cell(i, 4, _safe_text(p.get("note", "")))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
