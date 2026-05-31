"""报告方法声明测试（v2.9 A9）。"""
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from app.db.models import Project, FunctionPoint


def _seed_params(db):
    """Seed global params so get_effective resolves a complete cost context."""
    from app.services import params as ps
    ps.seed_from_csbmk(db=db)
    db.commit()


def _seed(db, pid, measurement_method="nesma_estimated"):
    p = Project(
        id=pid, name=f"report-test-{pid}",
        project_type="dev_only", phase="bidding",
        city="北京", industry="全行业",
        mode="forward", basis_data_ver="SSM-BK-202509",
        assessment_kind="development",
        measurement_method=measurement_method,
    )
    db.add(p)
    db.commit()
    return p


def _export_report(db, project_id, tmp_data_dir) -> Path:
    """Call generate_excel and return the output Path.

    generate_excel(db, project_id, band=None) → Path
    It writes to settings.export_dir / project_id / report_xxx.xlsx,
    so we need tmp_data_dir fixture to redirect export_dir to tmp_path.
    """
    from app.services.reports import generate_excel
    out = generate_excel(db, project_id)
    return Path(out)


def test_nesma_estimated_report_declaration(db_session, tmp_data_dir):
    """nesma_estimated 项目报告：声明含 NESMA。"""
    _seed_params(db_session)
    _seed(db_session, "p-a9-nesma", measurement_method="nesma_estimated")
    db_session.add(FunctionPoint(
        id="fp-a9-nesma-1", project_id="p-a9-nesma", version=1,
        category="EO", complexity="average", modify_type="add", ufp=5, us=5,
    ))
    db_session.commit()
    out = _export_report(db_session, "p-a9-nesma", tmp_data_dir)
    wb = load_workbook(str(out))
    found = False
    for ws_name in wb.sheetnames:
        for row in wb[ws_name].iter_rows(values_only=True):
            for cell in row:
                if cell and "NESMA" in str(cell):
                    found = True
    assert found, "报告中未找到 NESMA 声明"


def test_cosmic_report_declaration_and_conversion_note(db_session, tmp_data_dir):
    """cosmic 项目报告：声明含 COSMIC，并有 CFP→FP 换算备注。"""
    _seed_params(db_session)
    _seed(db_session, "p-a9-cosmic", measurement_method="cosmic")
    db_session.add(FunctionPoint(
        id="fp-a9-cosmic-1", project_id="p-a9-cosmic", version=1,
        category="EI", complexity="average", modify_type="add",
        ufp=8, us=8, cosmic_entry=2, cosmic_exit=2, cosmic_read=2, cosmic_write=2,
    ))
    db_session.commit()
    out = _export_report(db_session, "p-a9-cosmic", tmp_data_dir)
    wb = load_workbook(str(out))
    text_cells = []
    for ws_name in wb.sheetnames:
        for row in wb[ws_name].iter_rows(values_only=True):
            for cell in row:
                if cell:
                    text_cells.append(str(cell))
    full_text = " ".join(text_cells)
    assert "COSMIC" in full_text, "报告中未找到 COSMIC 声明"
    assert "换算" in full_text, "COSMIC 项目报告未包含换算备注"
