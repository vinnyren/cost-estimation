from io import BytesIO

from openpyxl import load_workbook

H = {"X-Auth-Token": "test-secret-token-xyz", "Origin": "http://127.0.0.1:8788"}


async def _make_project_with_fp(c) -> str:
    rp = await c.post("/api/projects", headers=H, json={
        "name": "T", "project_type": "dev_only", "phase": "bidding",
        "city": "北京", "industry": "电子政务", "mode": "forward",
        "basis_data_ver": "CSBMK®-202510",
    })
    pid = rp.json()["data"]["id"]
    await c.post(f"/api/projects/{pid}/functions",
                 headers={**H, "Content-Type": "application/json"},
                 json={"name": "首页", "category": "EQ", "complexity": "low",
                       "ufp": 4.0, "us": 4.0})
    return pid


async def test_download_excel(client_factory, tmp_data_dir):
    async with await client_factory() as c:
        pid = await _make_project_with_fp(c)
        r = await c.get(f"/api/reports/excel/{pid}", headers=H)
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
        wb = load_workbook(BytesIO(r.content))
        for s in ["封面", "评估结果汇总", "模块功能点及费用分项统计表",
                  "系统功能点明细表", "评估报告书", "调整因子表"]:
            assert s in wb.sheetnames
        # 评估结果汇总：调整后规模 S = us 4 × cf 1.25 = 5.0（在 D 列，SSM-BK-202509 招标阶段）
        ws = wb["评估结果汇总"]
        d_vals = [ws.cell(row=row, column=4).value
                  for row in range(1, ws.max_row + 1)]
        assert 5.0 in d_vals
        # 评估报告书必须有叙述内容（修复旧版空白 sheet）
        narr = wb["评估报告书"]
        narr_text = "\n".join(str(narr.cell(r, 1).value or "")
                              for r in range(1, narr.max_row + 1))
        assert "项目概述" in narr_text and "评估结论" in narr_text


async def test_reverse_report_total_matches_target(client_factory, tmp_data_dir):
    """反算项目的报告：总费用应复现目标造价（口径与结果页一致）。"""
    async with await client_factory() as c:
        rp = await c.post("/api/projects", headers=H, json={
            "name": "反算项目", "project_type": "dev_only", "phase": "bidding",
            "city": "北京", "industry": "电子政务", "mode": "reverse",
            "target_cost": 88.0,  # 万元
            "basis_data_ver": "CSBMK®-202510",
        })
        pid = rp.json()["data"]["id"]
        # 反算不依赖 FP 数值，但报告仍要求项目有 FP（FP 表作参考设计）
        await c.post(f"/api/projects/{pid}/functions",
                     headers={**H, "Content-Type": "application/json"},
                     json={"name": "首页", "category": "EQ", "complexity": "low",
                           "ufp": 4.0, "us": 4.0})
        r = await c.get(f"/api/reports/excel/{pid}", headers=H)
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        ws = wb["评估结果汇总"]
        # 汇总表 D 列（评估结果）里应出现评估总造价 ≈ 88 万元
        d_vals = [ws.cell(row=row, column=4).value
                  for row in range(1, ws.max_row + 1)]
        nums = [v for v in d_vals if isinstance(v, (int, float))]
        assert any(abs(v - 88.0) < 0.5 for v in nums), \
            f"评估总造价应 ≈ 88 万元，汇总 D 列实际为 {nums}"


async def test_download_no_fp_returns_400(client_factory, tmp_data_dir):
    async with await client_factory() as c:
        pid = await _make_project_with_fp(c)
        # 删掉刚加的 FP
        fps_r = await c.get(f"/api/projects/{pid}/functions", headers=H)
        fp_id = fps_r.json()["data"][0]["id"]
        await c.delete(f"/api/projects/{pid}/functions/{fp_id}", headers=H)
        r = await c.get(f"/api/reports/excel/{pid}", headers=H)
        assert r.status_code == 400
        assert r.json()["detail"]["error"]["code"] == "FP_EMPTY"
