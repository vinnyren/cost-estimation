from types import SimpleNamespace

from openpyxl import load_workbook

from app.exporters.excel import _safe_text
from app.exporters.report_builder import build_report

_SHEETS = ["封面", "评估结果汇总", "模块功能点及费用分项统计表",
           "系统功能点明细表", "评估报告书", "调整因子表"]


def _project():
    return SimpleNamespace(
        id="prj-test", name="测试项目", client="甲方", evaluator="乙方",
        phase="bidding", city="北京", industry="电子政务",
        basis_data_ver="CSBMK®-202510")


def _fp(**over):
    base = dict(subsystem="软件开发", l1_module="电子结算", l2_module="资金管理",
                name="客户账户查询", description="查询账户", category="ILF",
                ufp=35, us=35, reuse_level="低", modify_type="新增",
                source="manual", notes="")
    base.update(over)
    return SimpleNamespace(**base)


def _figures():
    return {
        "scale_us": 70.0,
        "scale_adjusted": 84.7,
        "cf_used": 1.21,
        "effort_dev": {"P10": 200.0, "P50": 540.0, "P90": 1300.0},
        "cost_dev": {"P10": 37000.0, "P50": 100000.0, "P90": 240000.0},
        "cost_ops": {"P10": 0.0, "P50": 0.0, "P90": 0.0},
        "cost_total": {"P10": 37000.0, "P50": 100000.0, "P90": 240000.0},
        "cost_total_p50_yuan": 100000.0,
        "dev_factor": 0.95,
        "rate_dev": 32198.0,
        "hours_per_pm": 174.0,
        "other_cost": 0.0,
    }


def test_build_report_creates_six_sheets(tmp_path):
    out = tmp_path / "r.xlsx"
    build_report(out, project=_project(),
                 functions=[_fp(), _fp(l1_module="智能终端", ufp=15, us=15)],
                 figures=_figures(), is_reverse=False, target_cost_wan=None)
    assert out.exists()
    wb = load_workbook(out)
    for s in _SHEETS:
        assert s in wb.sheetnames


def test_summary_has_total_cost(tmp_path):
    out = tmp_path / "r.xlsx"
    build_report(out, project=_project(), functions=[_fp()],
                 figures=_figures(), is_reverse=False, target_cost_wan=None)
    ws = load_workbook(out)["评估结果汇总"]
    # 评估结果列在 D 列；总造价（元）应出现 100000
    d_vals = [ws.cell(r, 4).value for r in range(1, ws.max_row + 1)]
    assert 100000.0 in d_vals


def test_narrative_sheet_not_empty(tmp_path):
    """评估报告书必须有实际叙述内容（修复旧版空白 sheet 的回归）。"""
    out = tmp_path / "r.xlsx"
    build_report(out, project=_project(), functions=[_fp()],
                 figures=_figures(), is_reverse=False, target_cost_wan=None)
    ws = load_workbook(out)["评估报告书"]
    text = "\n".join(str(ws.cell(r, 1).value or "")
                     for r in range(1, ws.max_row + 1))
    assert "项目概述" in text
    assert "评估结论" in text
    assert "测试项目" in text
    assert ws.max_row >= 10


def test_module_sheet_has_summary_row(tmp_path):
    out = tmp_path / "r.xlsx"
    build_report(out, project=_project(),
                 functions=[_fp(), _fp(l1_module="智能终端")],
                 figures=_figures(), is_reverse=False, target_cost_wan=None)
    ws = load_workbook(out)["模块功能点及费用分项统计表"]
    b_vals = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert "项目汇总" in b_vals


def test_reverse_report_shows_target(tmp_path):
    out = tmp_path / "r.xlsx"
    build_report(out, project=_project(), functions=[_fp()],
                 figures=_figures(), is_reverse=True, target_cost_wan=88.0)
    text = "\n".join(
        str(c.value or "")
        for row in load_workbook(out)["评估报告书"].iter_rows()
        for c in row)
    assert "反算" in text


def test_safe_text_quotes_formula():
    assert _safe_text("=cmd|calc") == "'=cmd|calc"
    assert _safe_text("用户登录") == "用户登录"
    assert _safe_text(42) == 42
    assert _safe_text(None) is None
